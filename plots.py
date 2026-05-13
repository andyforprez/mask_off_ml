import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter


# ── 1. Cutoff vs Player ───────────────────────────────────────────────────────

def _compute_actual_cutoff_series(df, cutoff_rank=24):
    df = df.copy()
    df['date'] = pd.to_datetime(df['date'])
    all_dates = sorted(df['date'].unique())

    last_pts = {}
    result   = []
    for d in all_dates:
        for _, row in df[df['date'] == d].iterrows():
            pid = row['player_id']
            last_pts[pid] = last_pts.get(pid, 0.0) + row['points']
        ranked     = sorted(last_pts.values(), reverse=True)
        cutoff_idx = cutoff_rank - 1
        cutoff     = ranked[cutoff_idx] if len(ranked) >= cutoff_rank else (ranked[-1] if ranked else 0)
        result.append({'date': d, 'cutoff': cutoff})
    return result


def plot_cutoff_vs_player(cutoff_history, expected_players, player_name,
                           player_path=None, df_actual=None, cutoff_rank=24):
    fig, axes = plt.subplots(1, 2, figsize=(15, 5))
    fig.suptitle(f'Playoff Projection — {player_name}', fontsize=13, fontweight='bold')

    ax = axes[0]

    # Actual historical cutoff (solid blue)
    if df_actual is not None:
        actual_cutoff = _compute_actual_cutoff_series(df_actual, cutoff_rank=cutoff_rank)
        ac_dates = pd.to_datetime([x['date'] for x in actual_cutoff])
        ac_vals  = [x['cutoff'] for x in actual_cutoff]
        ax.plot(ac_dates, ac_vals, color='steelblue', linewidth=2,
                label=f'{cutoff_rank}th place (actual)')

    # Projected cutoff: mean ± std (dashed blue)
    if cutoff_history:
        rows     = [{'date': pd.Timestamp(c['date']), 'cutoff': c['cutoff']}
                    for sim in cutoff_history for c in sim]
        cdf      = pd.DataFrame(rows)
        mean_cut = cdf.groupby('date')['cutoff'].mean()
        std_cut  = cdf.groupby('date')['cutoff'].std().fillna(0)

        stitch_date = ac_dates[-1]  if df_actual is not None else mean_cut.index[0]
        stitch_val  = ac_vals[-1]   if df_actual is not None else mean_cut.iloc[0]

        proj_dates = pd.to_datetime([stitch_date] + list(mean_cut.index))
        proj_vals  = np.array([stitch_val]  + list(mean_cut.values))
        proj_std   = np.array([0]           + list(std_cut.values))

        ax.plot(proj_dates, proj_vals, color='steelblue', linewidth=2,
                linestyle='--', label=f'{cutoff_rank}th place (projected mean)')
        ax.fill_between(proj_dates, proj_vals - proj_std, proj_vals + proj_std,
                        color='steelblue', alpha=0.15, label='±1 std')

    # Actual player history (solid red)
    last_actual_date = None
    last_actual_pts  = None
    if df_actual is not None:
        df2  = df_actual.copy()
        df2['date'] = pd.to_datetime(df2['date'])
        p_df = df2[df2['player_id'] == player_name].sort_values('date')
        if not p_df.empty:
            cum_pts = p_df['points'].cumsum()
            ax.plot(p_df['date'], cum_pts, color='crimson', linewidth=2,
                    label=f'{player_name} (actual)')
            last_actual_date = p_df['date'].iloc[-1]
            last_actual_pts  = float(cum_pts.iloc[-1])

    # Projected player path (dashed red)
    if player_path is not None and len(player_path) > 0:
        sim_dates = pd.to_datetime(list(player_path.index))
        sim_pts   = np.array(list(player_path.values))

        if last_actual_date is not None:
            stitch_d = pd.to_datetime([last_actual_date] + list(sim_dates))
            stitch_v = [last_actual_pts] + list(sim_pts)
        else:
            stitch_d, stitch_v = sim_dates, sim_pts

        ax.plot(stitch_d, stitch_v, color='crimson', linewidth=2,
                linestyle='--', label=f'{player_name} (projected mean)')

    ax.axvline(pd.Timestamp(df_actual['date'].max()) if df_actual is not None else pd.Timestamp.now(),
               color='gray', linewidth=1, linestyle=':', alpha=0.7, label='Today')

    ax.set_xlabel('Date')
    ax.set_ylabel('Cumulative Points')
    ax.legend(fontsize=8)
    ax.grid(axis='y', alpha=0.3)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %d'))
    ax.tick_params(axis='x', rotation=45)
    ax.set_title('Points Trajectory')

    # Panel 2: distribution of final cutoff
    ax2 = axes[1]
    if cutoff_history:
        final_cutoffs = np.array([sim[-1]['cutoff'] for sim in cutoff_history if sim])
        mu = float(np.mean(final_cutoffs))
        sd = float(np.std(final_cutoffs))

        ax2.hist(final_cutoffs, bins=40, color='steelblue', edgecolor='white',
                 linewidth=0.4, alpha=0.85, label='Cutoff distribution')
        ax2.axvline(mu, color='navy', linewidth=2, label=f'Mean: {mu:.0f}')
        ax2.axvline(mu - sd, color='navy', linewidth=1, linestyle='--', alpha=0.6)
        ax2.axvline(mu + sd, color='navy', linewidth=1, linestyle='--', alpha=0.6,
                    label=f'±1σ: {sd:.0f}')

        if player_path is not None and len(player_path) > 0:
            proj_final = float(player_path.values[-1])
            if last_actual_pts is not None:
                proj_final += last_actual_pts
            ax2.axvline(proj_final, color='crimson', linewidth=2.5,
                        label=f'{player_name}: {proj_final:.0f}')
            pct = float(np.mean(final_cutoffs <= proj_final) * 100)
            ax2.set_title(f'Final Cutoff Distribution\n'
                          f'{player_name} clears cutoff in {pct:.1f}% of sims')
        else:
            ax2.set_title('Final Cutoff Distribution')

        ax2.set_xlabel(f'{cutoff_rank}th-place Points')
        ax2.set_ylabel('Frequency (simulations)')
        ax2.legend(fontsize=8)
        ax2.grid(axis='y', alpha=0.3)

    plt.tight_layout()
    plt.savefig('data/cutoff_vs_player.png', dpi=150, bbox_inches='tight')
    plt.show()
    print("Saved data/cutoff_vs_player.png")


# ── 2. Rank Projections Multi ─────────────────────────────────────────────────

def plot_rank_projections_multi(real_ranks_dict, sim_ranks_dict, today,
                                 top_n=24, cutoff_rank=24):
    players = list(real_ranks_dict.keys())[:top_n]

    fig, ax = plt.subplots(figsize=(14, 7))
    cmap    = plt.get_cmap('tab20')
    colors  = {p: cmap(i % 20) for i, p in enumerate(players)}

    for player in players:
        color = colors[player]

        real    = real_ranks_dict.get(player, [])
        r_dates = pd.to_datetime([x['date'] for x in real if x.get('rank') is not None])
        r_ranks = [x['rank'] for x in real if x.get('rank') is not None]

        if len(r_dates):
            ax.plot(r_dates, r_ranks, color=color, linewidth=1.6, alpha=0.9)

        sim = sim_ranks_dict.get(player)
        if sim is not None and len(sim) > 0:
            s_dates = pd.to_datetime(list(sim.index))
            s_ranks = list(sim.values)

            if len(r_dates):
                s_dates = pd.to_datetime([r_dates[-1]] + list(s_dates))
                s_ranks = [r_ranks[-1]] + s_ranks

            ax.plot(s_dates, s_ranks, color=color, linewidth=1.6,
                    linestyle='--', alpha=0.9, label=player)

            if len(s_dates):
                ax.annotate(player,
                            xy=(s_dates[-1], s_ranks[-1]),
                            xytext=(4, 0), textcoords='offset points',
                            fontsize=5.5, color=color, va='center', clip_on=True)

    ax.axvline(pd.Timestamp(today), color='black', linewidth=1.5,
               linestyle=':', alpha=0.6, label='Today')
    ax.axhline(cutoff_rank + 0.5, color='limegreen', linewidth=1.5,
               linestyle='--', alpha=0.7, label=f'Playoff cutoff ({cutoff_rank})')

    ax.invert_yaxis()
    ax.set_ylim(top_n + 2, 0)
    ax.set_xlabel('Date')
    ax.set_ylabel('Rank')
    ax.set_title(f'Rank Projections — Top {top_n} Players  (solid=actual, dashed=projected)')
    ax.legend(fontsize=5.5, ncol=3, loc='lower left')
    ax.grid(axis='y', alpha=0.25)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %d'))
    ax.tick_params(axis='x', rotation=45)

    plt.tight_layout()
    plt.savefig('data/rank_projections.png', dpi=150, bbox_inches='tight')
    plt.show()
    print("Saved data/rank_projections.png")


# ── 3. Playoff Odds Excel ─────────────────────────────────────────────────────

def save_playoff_odds_excel(df, path='data/playoff_odds.xlsx', cutoff_rank=24):
    """
    Playoff-odds workbook. cutoff_rank drives which rank columns to include
    and where the red border goes.
    """
    df = df.copy()
    top_col = [c for c in df.columns if c.startswith('Top')][0]

    # Grab rank columns up to cutoff_rank
    rank_cols = sorted(
        [c for c in df.columns
         if c.startswith('Rank ') and int(c.split()[1]) <= cutoff_rank],
        key=lambda c: int(c.split()[1])
    )
    df = df[[top_col] + rank_cols].copy()

    # Expected rank (weighted average)
    df['_exp'] = sum(
        df[f'Rank {i}'] * i for i in range(1, cutoff_rank + 1)
        if f'Rank {i}' in df.columns
    )
    df = df.sort_values(by=[top_col, '_exp'], ascending=[False, True]).drop(columns=['_exp'])
    df.insert(0, 'Final Rank', range(1, len(df) + 1))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Playoff Odds'

    headers = ['Player', 'Final Rank', f'Top {cutoff_rank} %'] + \
              [f'#{i}' for i in range(1, cutoff_rank + 1)]
    ws.append(headers)

    for player, row in df.iterrows():
        ws.append([player] + list(row.values))

    n_rows = len(df)
    n_cols = len(headers)

    # Header styling
    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Format floats as percentages; zero → blank
    for row in ws.iter_rows(min_row=2, min_col=3, max_row=n_rows + 1, max_col=n_cols):
        for cell in row:
            if isinstance(cell.value, float):
                if cell.value == 0.0:
                    cell.value = None
                else:
                    cell.number_format = '0.0%'

    # Green gradient on Top-N % column (C)
    ws.conditional_formatting.add(f'C2:C{n_rows + 1}', ColorScaleRule(
        start_type='num', start_value=0,   start_color='FF4444',
        mid_type='num',   mid_value=0.5,   mid_color='FFFF00',
        end_type='num',   end_value=1,     end_color='00CC44',
    ))

    # Blue gradient on per-rank columns (D+)
    last_col_letter = get_column_letter(n_cols)
    ws.conditional_formatting.add(f'D2:{last_col_letter}{n_rows + 1}', ColorScaleRule(
        start_type='min',       start_color='FFFFFF',
        mid_type='percentile',  mid_value=80, mid_color='9EC5FE',
        end_type='max',         end_color='0D6EFD',
    ))

    # Light green fill for Final Rank ≤ cutoff_rank
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    ws.conditional_formatting.add(f'B2:B{n_rows + 1}', CellIsRule(
        operator='lessThanOrEqual', formula=[str(cutoff_rank)], fill=green_fill
    ))

    # Red border below playoff cutoff row
    red_side     = Side(style='medium', color='FF0000')
    border_row   = cutoff_rank + 1   # +1 for header row
    for col in range(1, n_cols + 1):
        ws.cell(row=border_row, column=col).border = Border(bottom=red_side)

    # Alternating row shading
    alt_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    for r in range(2, n_rows + 2, 2):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).fill = alt_fill

    # Column widths
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    for col_idx in range(4, n_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 7

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'D2'

    wb.save(path)
    print(f"Saved {path}")